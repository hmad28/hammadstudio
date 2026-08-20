"""Build the Hammad Studio keyword-to-URL execution workbook.

This script intentionally maps query variants to canonical intent pages. It does
not generate one page per keyword and it marks unsupported services/locations as
hold items instead of turning them into thin or misleading pages.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


SITE = "https://hammad.studio"
DEFAULT_MASTER = Path(r"C:\Users\Pongo\Downloads\Hammad-Studio-SEO-AEO-GEO-Keyword-Master-2026.xlsx")
DEFAULT_SUPPLEMENT = Path(r"C:\Users\Pongo\Downloads\Keyword_SEO_AEO_GEO_HammadStudio_500plus.xlsx")
DEFAULT_OUTPUT = Path("docs/Hammad-Studio-SEO-AEO-GEO-Implementation-Map-2026.xlsx")

PURPLE = "8057FF"
LIME = "D2F34C"
NAVY = "070A12"
OFF_WHITE = "F7F5EF"
LIGHT_PURPLE = "EEE9FF"
LIGHT_GREEN = "EAF5D0"
LIGHT_YELLOW = "FFF1C2"
LIGHT_RED = "FCE2E2"
GRAY = "E5E7EB"
WHITE = "FFFFFF"


PAGE_BLUEPRINTS = [
    ("/", "Core website + entity", "jasa pembuatan website", "Jasa Pembuatan Website Profesional | Hammad Studio", "Jasa pembuatan website untuk UMKM dan perusahaan. Website modern, cepat, mobile-friendly, SEO-ready, dan sesuai kebutuhan bisnis.", "Website profesional untuk bisnis yang ingin berkembang lebih serius.", "Organization; ProfessionalService; WebSite", "/jasa-pembuatan-website; /harga-website; /work; /studio", "Live — improve entity and proof"),
    ("/jasa-pembuatan-website", "Core transactional", "jasa pembuatan website profesional", "Jasa Pembuatan Website Profesional | Hammad Studio", "Jasa pembuatan website profesional untuk bisnis dengan desain, performa, SEO foundation, analytics, dan engineering yang rapi.", "Website profesional yang dibangun untuk dipercaya dan menghasilkan peluang.", "Service; FAQPage; BreadcrumbList", "/jasa-landing-page; /jasa-website-umkm; /jasa-website-company-profile; /work", "Live — canonical money page"),
    ("/harga-website", "Pricing", "harga pembuatan website", "Harga Pembuatan Website 2026 | Hammad Studio", "Harga website mulai Rp499 ribu. Bandingkan Website Starter, CMS, e-commerce, booking, POS, CRM, LMS, dan custom software.", "Waktunya bisnis kamu ikut merdeka secara digital.", "Service; OfferCatalog", "/jasa-pembuatan-website; /faq; /contact", "Live — canonical pricing page"),
    ("/jasa-landing-page", "Landing page", "jasa pembuatan landing page", "Jasa Pembuatan Landing Page untuk Bisnis | Hammad Studio", "Landing page untuk iklan, campaign, produk, dan lead generation dengan WhatsApp, form, analytics, performa, dan SEO foundation.", "Satu halaman. Satu pesan. Satu tindakan yang jelas.", "Service; FAQPage; BreadcrumbList", "/harga-website; /jasa-website-umkm; /work/naiklevel-ai", "Implemented 2026-08-20"),
    ("/jasa-website-company-profile", "Company profile", "jasa website company profile", "Jasa Website Company Profile Perusahaan | Hammad Studio", "Website company profile untuk perusahaan yang membutuhkan kredibilitas, layanan jelas, bukti project, inquiry, dan SEO foundation.", "Presentasi digital perusahaan yang layak dipercaya.", "Service; FAQPage; BreadcrumbList", "/redesign-website; /maintenance-website; /work/add-logistik", "Live — canonical company page"),
    ("/jasa-website-umkm", "UMKM", "jasa pembuatan website UMKM", "Jasa Pembuatan Website UMKM | Hammad Studio", "Website UMKM untuk profil, produk, katalog, WhatsApp, Maps, dan fondasi pencarian Google dengan scope yang transparan.", "Bantu bisnis kecil terlihat serius, jelas, dan siap tumbuh.", "Service; FAQPage; BreadcrumbList", "/jasa-landing-page; /services/e-commerce; /work/shofi-frozen", "Live — canonical UMKM page"),
    ("/services/e-commerce", "E-commerce", "jasa pembuatan website e-commerce", "Jasa Pembuatan Website E-Commerce | Hammad Studio", "Toko online dengan katalog, cart, checkout, QRIS atau full payment, order management, inventory, dan dashboard.", "Bangun toko online yang terasa milik brand Anda sendiri.", "Service; BreadcrumbList", "/harga-website; /solusi/umkm/toko-online; /work/cpx-jersey", "Live — canonical commerce page"),
    ("/jasa-web-app", "Web app + business apps", "jasa pembuatan web app", "Jasa Pembuatan Web App dan Dashboard Custom | Hammad Studio", "Web app, dashboard, booking, CRM, POS, inventory, portal, dan sistem bisnis dengan role, database, workflow, serta API.", "Saat bisnis membutuhkan fungsi, data, dan workflow—bukan halaman saja.", "Service; FAQPage; BreadcrumbList", "/website-custom; /harga-website#business-apps; /work/studyshare", "Implemented 2026-08-20"),
    ("/website-custom", "Custom software", "jasa pembuatan website custom", "Jasa Pembuatan Website dan Software Custom | Hammad Studio", "Website, portal, SaaS, dashboard, integrasi, dan custom software yang dirancang berdasarkan workflow serta kebutuhan bisnis.", "Solusi digital yang mengikuti workflow bisnis Anda.", "Service; FAQPage; BreadcrumbList", "/jasa-web-app; /solusi/perusahaan/sistem-custom-dashboard; /contact", "Live — canonical custom page"),
    ("/redesign-website", "Redesign", "jasa redesign website", "Jasa Redesign Website Perusahaan | Hammad Studio", "Redesign website untuk memperbaiki kredibilitas, struktur, performa, SEO foundation, UX mobile, dan jalur inquiry.", "Ubah website lama menjadi aset bisnis yang kembali relevan.", "Service; FAQPage; BreadcrumbList", "/maintenance-website; /jasa-website-company-profile", "Live — canonical redesign page"),
    ("/maintenance-website", "Maintenance", "jasa maintenance website", "Jasa Maintenance dan Perawatan Website | Hammad Studio", "Maintenance website untuk update, monitoring, backup, bug fixing, performa, keamanan, konten, dan support berkelanjutan.", "Website tetap sehat setelah hari peluncuran.", "Service; FAQPage; BreadcrumbList", "/redesign-website; /insight/checklist-keamanan-website-perusahaan", "Implemented 2026-08-20"),
    ("/industri/travel", "Travel / Umrah / Haji", "jasa website travel umrah", "Jasa Website Travel Umrah dan Haji | Hammad Studio", "Website travel dengan paket, jadwal, legal information, galeri, edukasi, konsultasi, pendaftaran, dan CMS.", "Bangun kepercayaan calon jamaah sebelum percakapan pertama.", "Service; FAQPage; BreadcrumbList", "/work/jamwisata; /work/sahabat-qolbu; /work/nusuk-haromain-indonesia", "Live — proof-supported vertical"),
    ("/industri/logistik", "Logistics", "jasa website perusahaan logistik", "Jasa Website Perusahaan Logistik | Hammad Studio", "Website logistik untuk layanan, armada, coverage, RFQ, kapabilitas operasional, dan inquiry perusahaan.", "Kapabilitas logistik yang mudah dinilai calon klien B2B.", "Service; FAQPage; BreadcrumbList", "/work/add-logistik; /work/buraq-logistik", "Live — proof-supported vertical"),
    ("/industri/umkm-kuliner", "Culinary", "jasa website kuliner", "Jasa Website UMKM Kuliner | Hammad Studio", "Website kuliner dengan menu, katalog, promo, lokasi, WhatsApp order, dan pengelolaan konten.", "Buat produk kuliner lebih mudah ditemukan, dipilih, dan dipesan.", "Service; FAQPage; BreadcrumbList", "/work/shofi-frozen; /jasa-website-umkm", "Live — proof-supported vertical"),
    ("/industri/pendidikan", "Education", "jasa website pendidikan", "Jasa Website dan Platform Pendidikan | Hammad Studio", "Website sekolah, lembaga pendidikan, course, dan platform belajar dengan CMS, registrasi, akun, materi, serta dashboard.", "Informasi pendidikan yang jelas, platform belajar yang mudah digunakan.", "Service; FAQPage; BreadcrumbList", "/work/studyshare; /jasa-web-app", "Live — proof-supported vertical"),
    ("/industri/fashion", "Fashion", "jasa website fashion", "Jasa Website Fashion dan Apparel | Hammad Studio", "Website fashion dengan katalog koleksi, varian, lookbook, custom order, marketplace, dan commerce flow.", "Tampilkan koleksi dan karakter brand tanpa batasan marketplace.", "Service; FAQPage; BreadcrumbList", "/work/cpx-jersey; /services/e-commerce", "Live — proof-supported vertical"),
    ("/industri/event", "Event", "jasa website event", "Jasa Website dan Ticketing Event | Hammad Studio", "Website event, registrasi peserta, pembayaran, tiket digital, QR scanner, check-in, dan dashboard attendance.", "Promosi, registrasi, tiket, dan check-in dalam alur yang jelas.", "Service; FAQPage; BreadcrumbList", "/jasa-web-app; /harga-website#plan-ticketing-system", "Live — product-supported vertical"),
    ("/jasa-pembuatan-website-jakarta", "Jakarta local", "jasa pembuatan website Jakarta", "Jasa Pembuatan Website Jakarta | Hammad Studio", "Jasa pembuatan website berbasis Jakarta untuk UMKM dan perusahaan, dengan proses konsultasi serta delivery online.", "Partner website berbasis Jakarta untuk bisnis yang ingin tumbuh serius.", "Service; FAQPage; BreadcrumbList", "/contact; /jasa-pembuatan-website; /work", "Live — only verified local page"),
    ("/insight", "Knowledge hub", "panduan website bisnis", "Insight Website untuk Bisnis | Hammad Studio", "Panduan harga, scope, company profile, website UMKM, teknologi, SEO, AEO, GEO, dan pengembangan website bisnis.", "Panduan website untuk keputusan bisnis yang lebih jelas.", "CollectionPage; ItemList", "/jasa-pembuatan-website; /harga-website", "Live — answer content hub"),
    ("/work", "Proof / case studies", "portfolio Hammad Studio", "Karya dan Case Study Website | Hammad Studio", "Portfolio website, e-commerce, travel, logistik, pendidikan, platform, dan custom development Hammad Studio.", "Project nyata. Masalah nyata. Solusi yang bisa dilihat.", "CollectionPage; ItemList", "/jasa-pembuatan-website; /contact", "Live — proof hub"),
    ("/studio", "Entity / E-E-A-T", "tentang Hammad Studio", "Tentang Hammad Studio | Web Development Jakarta", "Profil Hammad Studio, cara kerja, founder, standar engineering, fokus layanan, dan project website dari Jakarta untuk Indonesia.", "Studio independen untuk website dan produk digital.", "AboutPage; Organization; Person; BreadcrumbList", "/work; /contact; /jasa-pembuatan-website", "Live — entity page"),
    ("/faq", "AEO / trust", "FAQ jasa pembuatan website", "FAQ Jasa Pembuatan Website | Hammad Studio", "Jawaban tentang harga, domain, hosting, CMS, timeline, revisi, SEO, support, ownership, legalitas, dan cara memulai project.", "Pertanyaan yang biasanya muncul sebelum project dimulai.", "FAQPage; BreadcrumbList", "/harga-website; /contact", "Live — answer hub"),
]


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def map_keyword(keyword: str, cluster: str = "", recommended: str = "") -> tuple[str, str, str]:
    k = keyword.lower()
    c = cluster.lower()
    rec = recommended.lower()

    unsupported = ("branding", "brand identity", "logo", "video", "videografi", "fotografi", "photo studio", "motion graphic", "creative production", "packaging", "social media content", "ui ux")
    if any(term in k or term in c for term in unsupported):
        return "", "Hold — unsupported offering", "Do not publish until Hammad Studio offers and can prove this service."

    other_cities = ("bandung", "surabaya", "yogyakarta", "jogja", "semarang", "medan", "makassar", "bali", "bogor", "depok", "tangerang", "bekasi", "banten")
    if any(city in k for city in other_cities):
        return "/jasa-pembuatan-website", "Hold — no verified local presence", "Do not create a doorway city page. Serve remotely from the canonical national service page."
    if "jakarta" in k or "local / service area" in c:
        return "/jasa-pembuatan-website-jakarta", "Implement on verified local page", "Only Jakarta is a verified location; Indonesia remains the remote service area."

    if "harga" in k or "biaya" in k or "paket" in k or "pricing" in c:
        return "/harga-website", "Implement on canonical pricing page", "Cover variants naturally in pricing, comparison, and FAQ content."
    if "landing page" in k or "landing page" in c:
        return "/jasa-landing-page", "Implement on canonical service page", "One landing-page intent page; no exact-match duplicates."
    if "company profile" in k or "company profile" in c or "website perusahaan" in k or "web perusahaan" in k:
        return "/jasa-website-company-profile", "Implement on canonical service page", "Company profile and company website variants share one commercial intent."
    if "umkm" in k or "umkm" in c:
        return "/jasa-website-umkm", "Implement on canonical service page", "Use business outcomes, WhatsApp, catalog, Maps, and proof."
    if any(term in k for term in ("e-commerce", "ecommerce", "toko online", "online store", "qris", "payment gateway", "checkout")) or "e-commerce" in c:
        return "/services/e-commerce", "Implement on canonical commerce page", "Cover QRIS and full-payment variants without separate thin pages."
    if "maintenance" in k or "perawatan website" in k:
        return "/maintenance-website", "Implement on canonical service page", "Maintenance has a dedicated scope and commercial intent."
    if any(term in k for term in ("redesign", "redesain", "revamp", "renovasi website", "migrasi website")):
        return "/redesign-website", "Implement on canonical service page", "Redesign and revamp variants share one page."
    if any(term in k for term in ("crm", "pos system", "inventory", "booking system", "ticketing", "membership", "lms", "course platform", "rental system", "dashboard", "web app", "aplikasi web", "sistem informasi web")):
        return "/jasa-web-app", "Implement on canonical web-app page", "Specific business-app terms are sections/product anchors, not separate thin pages yet."
    if any(term in k for term in ("custom software", "software development", "saas", "erp", "marketplace", "multi-tenant", "custom web")) or "business systems" in c:
        return "/website-custom", "Implement on canonical custom page", "Custom software intent requires discovery, proof, and architecture context."
    if any(term in k for term in ("travel", "umrah", "haji")) or "travel" in c:
        return "/industri/travel", "Implement on proof-supported industry page", "Supported by JamWisata, Sahabat Qolbu, and Nusuk Haromain."
    if "logistik" in k or "ekspedisi" in k or "logistics" in k or "logistik" in c:
        return "/industri/logistik", "Implement on proof-supported industry page", "Supported by ADD Logistik and Buraq Logistik."
    if any(term in k for term in ("kuliner", "restoran", "restaurant", "frozen food", "catering", "bakery", "makanan")) or "kuliner" in c:
        return "/industri/umkm-kuliner", "Implement on proof-supported industry page", "Supported by Shofi Frozen and commerce capabilities."
    if any(term in k for term in ("sekolah", "pendidikan", "education", "course", "kampus", "bimbel")) or "education" in c:
        return "/industri/pendidikan", "Implement on proof-supported industry page", "Supported by StudyShare and education platform capabilities."
    if any(term in k for term in ("fashion", "jersey", "apparel")) or "fashion" in c:
        return "/industri/fashion", "Implement on proof-supported industry page", "Supported by CPX Jersey."
    if "event" in k or "ticket" in k or "event" in c:
        return "/industri/event", "Implement on product-supported industry page", "Use ticketing capability without inventing event outcomes."
    if any(term in k for term in ("rental mobil", "rental motor", "automotive")) or "automotive" in c:
        return "/jasa-web-app", "Implement on web-app page with rental proof", "Supported by Drivemate; avoid a thin automotive vertical for now."
    if any(term in k for term in ("masjid", "pesantren", "yayasan islam", "islamic")) or "islamic" in c:
        return "/jasa-website-company-profile", "Consolidate until a full industry page is justified", "Use Masjid Raya Puri Telukjambe as proof; do not mass-produce niche pages."
    if any(term in k for term in ("property", "properti", "hotel", "villa", "healthcare", "klinik", "rumah sakit", "manufaktur", "konstruksi", "distributor", "ekspor", "impor")):
        return "/jasa-website-company-profile", "Hold industry page — evidence required", "Map to the canonical company service until genuine first-party proof is available."
    if any(term in c for term in ("aeo", "geo", "informational")) or k.startswith(("apa ", "bagaimana ", "kenapa ", "berapa ", "cara ", "rekomendasi ", "siapa ")):
        if "ai" in k or "geo" in k or "aeo" in k or "ai overview" in k or "chatgpt" in k:
            return "/insight/seo-aeo-geo-untuk-website-bisnis", "Implement as answer-first content", "Answer clearly; avoid special-AI-schema claims."
        return "/insight", "Content backlog — consolidate by topic", "Group related questions into one useful article, then link to the relevant money page."
    if "brand" in c or "hammad studio" in k or "hammad.studio" in k or "geo / ai recommendation" in c:
        return "/", "Implement through entity clarity", "Use consistent entity facts, sameAs, proof, pricing, and contact—not forced query text."
    if "core website" in c or "jasa pembuatan website" in k or "jasa website" in k or "web developer" in k or "studio pembuatan website" in k:
        return "/jasa-pembuatan-website", "Implement on canonical core page", "Use natural semantic coverage and proof."
    if rec:
        if rec.startswith("/"):
            return rec, "Review recommended canonical URL", "Use the source recommendation after checking intent and evidence."
        if "homepage" in rec:
            return "/", "Implement through entity and service context", "Avoid stuffing the homepage with every modifier."
        if "insight" in rec:
            return "/insight", "Content backlog — consolidate by topic", "Create only when the topic adds unique value."
    return "/jasa-pembuatan-website", "Consolidate on core page", "No separate page until intent, proof, and business value justify it."


def rows_from_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, str]]]:
    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book[sheet_name]
    values = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(values)]
    rows = []
    for values_row in values:
        row = {headers[index]: clean(value) for index, value in enumerate(values_row)}
        if any(row.values()):
            rows.append(row)
    book.close()
    return headers, rows


def add_table(sheet, name: str, end_row: int, end_col: int) -> None:
    if end_row < 2:
        return
    from openpyxl.utils import get_column_letter
    table = Table(displayName=name, ref=f"A1:{get_column_letter(end_col)}{end_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)


def style_sheet(sheet, widths: dict[str, float] | None = None) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=9, color="1F2937")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if widths:
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width


def write_rows(sheet, headers: Iterable[str], rows: Iterable[Iterable[object]]) -> None:
    sheet.append(list(headers))
    for row in rows:
        sheet.append(list(row))


def build_workbook(master_path: Path, supplement_path: Path, output_path: Path) -> None:
    _, master_rows = rows_from_sheet(master_path, "Keyword Master")
    _, priority_rows = rows_from_sheet(master_path, "Priority Keywords")
    _, supplement_rows = rows_from_sheet(supplement_path, "Keyword List 500+")

    workbook = Workbook()
    workbook.remove(workbook.active)

    overview = workbook.create_sheet("README")
    overview.sheet_view.showGridLines = False
    overview.merge_cells("A1:F1")
    overview["A1"] = "HAMMAD STUDIO — SEO / AEO / GEO IMPLEMENTATION MAP 2026"
    overview["A1"].fill = PatternFill("solid", fgColor=NAVY)
    overview["A1"].font = Font(name="Arial", size=18, bold=True, color=WHITE)
    overview["A1"].alignment = Alignment(vertical="center")
    overview.row_dimensions[1].height = 42
    summary = [
        ("Purpose", "Turn keyword research into canonical pages, answer content, schema, and internal links."),
        ("Core rule", "One search intent = one strong canonical page. Never create a page for every exact-match keyword."),
        ("Entity", "Hammad Studio → Jakarta-based website development studio → Indonesia/remote → website, commerce, web app, business systems, custom software."),
        ("Local rule", "Only Jakarta receives a local landing page. Other city variants are held until a genuine presence exists."),
        ("Evidence rule", "Industry pages require real projects or first-party expertise. Unsupported branding/video/UI-UX keywords are held."),
        ("AEO rule", "Use a direct answer, then evidence, examples, limitations, and a next step."),
        ("GEO rule", "No special AI schema. Use crawlable content, consistent entity facts, relevant structured data, proof, authorship, citations, and reputation."),
        ("Sources", f"{master_path.name}; {supplement_path.name}"),
        ("Generated", "2026-08-20 — Asia/Jakarta"),
    ]
    for index, (label, value) in enumerate(summary, start=3):
        overview.cell(index, 1, label).font = Font(name="Arial", bold=True, color=PURPLE)
        overview.cell(index, 2, value).font = Font(name="Arial", size=10, color="1F2937")
        overview.cell(index, 2).alignment = Alignment(wrap_text=True, vertical="top")
        overview.merge_cells(start_row=index, start_column=2, end_row=index, end_column=6)
        overview.row_dimensions[index].height = 34
    for column, width in {"A": 22, "B": 28, "C": 22, "D": 22, "E": 22, "F": 22}.items():
        overview.column_dimensions[column].width = width

    blueprints = workbook.create_sheet("Page Blueprints")
    blueprint_headers = ["Canonical URL", "Intent Cluster", "Primary Keyword", "Meta Title", "Title Chars", "Meta Description", "Description Chars", "H1", "Schema", "Internal Links", "Status"]
    blueprint_rows = [(url, cluster, primary, title, len(title), description, len(description), h1, schema, links, status) for url, cluster, primary, title, description, h1, schema, links, status in PAGE_BLUEPRINTS]
    write_rows(blueprints, blueprint_headers, blueprint_rows)
    style_sheet(blueprints, {"A": 34, "B": 24, "C": 34, "D": 54, "E": 12, "F": 64, "G": 16, "H": 52, "I": 32, "J": 54, "K": 30})
    add_table(blueprints, "PageBlueprints", blueprints.max_row, blueprints.max_column)
    for row in range(2, blueprints.max_row + 1):
        blueprints.cell(row, 5).fill = PatternFill("solid", fgColor=LIGHT_GREEN if blueprints.cell(row, 5).value <= 60 else LIGHT_YELLOW)
        blueprints.cell(row, 7).fill = PatternFill("solid", fgColor=LIGHT_GREEN if blueprints.cell(row, 7).value <= 160 else LIGHT_YELLOW)

    mapping = workbook.create_sheet("Keyword Mapping")
    mapping_headers = ["Source", "Keyword", "Cluster / Category", "Intent", "Funnel", "Priority", "Source Recommended URL", "Canonical URL", "Absolute URL", "Decision", "Implementation Note"]
    mapping_rows = []
    for row in master_rows:
        url, decision, note = map_keyword(row.get("Keyword", ""), row.get("Cluster", ""), row.get("Target / Recommended URL", ""))
        mapping_rows.append(("Keyword Master", row.get("Keyword"), row.get("Cluster"), row.get("Intent"), row.get("Funnel"), row.get("Priority"), row.get("Target / Recommended URL"), url, f"{SITE}{url}" if url else "", decision, note))
    for row in supplement_rows:
        url, decision, note = map_keyword(row.get("Keyword", ""), row.get("Kategori", ""), "")
        mapping_rows.append(("500+ Supplemental", row.get("Keyword"), row.get("Kategori"), row.get("Intent"), "", row.get("Prioritas"), "", url, f"{SITE}{url}" if url else "", decision, note))
    write_rows(mapping, mapping_headers, mapping_rows)
    style_sheet(mapping, {"A": 20, "B": 48, "C": 34, "D": 18, "E": 14, "F": 14, "G": 34, "H": 38, "I": 52, "J": 38, "K": 70})
    add_table(mapping, "KeywordMapping", mapping.max_row, mapping.max_column)
    for row in range(2, mapping.max_row + 1):
        decision = clean(mapping.cell(row, 10).value)
        color = LIGHT_RED if decision.startswith("Hold") else LIGHT_YELLOW if "backlog" in decision.lower() or "review" in decision.lower() else LIGHT_GREEN
        mapping.cell(row, 10).fill = PatternFill("solid", fgColor=color)

    priority = workbook.create_sheet("Priority Execution")
    priority_headers = ["Keyword", "Cluster", "Intent", "Funnel", "Priority", "Canonical URL", "Decision", "Why / Next Action"]
    priority_data = []
    for row in priority_rows:
        url, decision, note = map_keyword(row.get("Keyword", ""), row.get("Cluster", ""), row.get("Target / Recommended URL", ""))
        priority_data.append((row.get("Keyword"), row.get("Cluster"), row.get("Intent"), row.get("Funnel"), row.get("Priority"), url, decision, note))
    write_rows(priority, priority_headers, priority_data)
    style_sheet(priority, {"A": 50, "B": 38, "C": 18, "D": 14, "E": 14, "F": 40, "G": 40, "H": 72})
    add_table(priority, "PriorityExecution", priority.max_row, priority.max_column)

    guardrails = workbook.create_sheet("Implementation Rules")
    rules = [
        ("Canonical intent", "One strong URL covers natural language variants with the same intent."),
        ("Keyword use", "Primary query in title/H1 where natural; semantic variants in H2, body, FAQ, anchors, and examples."),
        ("No stuffing", "Do not paste keyword lists into visible copy, metadata, alt text, or schema."),
        ("No doorway pages", "Do not create city pages unless Hammad Studio has genuine local operations and unique local value."),
        ("No fake proof", "Do not fabricate volume, ranking, revenue, speed, testimonials, clients, certifications, reviews, or outcomes."),
        ("AEO", "Start with a self-contained answer; follow with evidence, comparison, steps, caveats, and CTA."),
        ("GEO", "Keep entity facts consistent; strengthen project proof, authorship, first-party data, and legitimate external citations."),
        ("Schema", "Use Organization/WebSite, Service, FAQPage, Article, BreadcrumbList, CollectionPage, and OfferCatalog only when their visible content exists."),
        ("Internal links", "Use descriptive anchors from insight → service, service → proof, proof → service, and pricing → relevant service."),
        ("Content threshold", "Publish a new page only when it has distinct intent, unique value, sufficient evidence, and a real conversion path."),
        ("Measurement", "Track indexed pages, non-brand clicks, money-keyword impressions, WhatsApp/form conversions, qualified leads, deals, and organic revenue."),
    ]
    write_rows(guardrails, ["Rule", "Implementation Standard"], rules)
    style_sheet(guardrails, {"A": 28, "B": 110})
    add_table(guardrails, "ImplementationRules", guardrails.max_row, guardrails.max_column)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--master", type=Path, default=DEFAULT_MASTER)
    parser.add_argument("--supplement", type=Path, default=DEFAULT_SUPPLEMENT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    build_workbook(args.master, args.supplement, args.output)
    print(args.output.resolve())


if __name__ == "__main__":
    main()
